import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, colors
from openpyxl.utils import get_column_letter
import re

# 設定要讀取的檔案路徑和開頭字串
file_path = './DDQN-6-variable-20230602-055743/models/tested_reward_state'
file_prefix = 'tested_reward_state_'  # 或者是其他開頭字串 # 1-A-10

# 獲取符合開頭字串的所有檔案路徑
file_list = [os.path.join(file_path, f) for f in os.listdir(file_path) if f.startswith(file_prefix)]
parts_list = [file_name.split('/')[-1].split('.')[0].split('_')[-1] for file_name in file_list]

df_list = [pd.read_excel(f,header=None, \
                   names=['torque', 'reachable', 'manipulator', 'axis2', 'axis3', 'torque_sum'\
                           ,'nan', 'reward', 'nan', 'motor2', 'motor3']) for f in file_list]

total_reachable_diff = 0
total_manipulator_diff = 0

total_torque_diff = 0

total_reachable_ori = 0
total_manipulator_ori = 0

total_torque_ori = 0

total_reachable_best = 0
total_manipulator_best = 0

total_torque_best = 0

optimal_design = Workbook()
sheet_optimal_design = optimal_design.active
i = 0

count_a = 0
count_b = 0
for _ in range(len(file_list)):
        df_sorted = df_list[_].sort_values(by=['reward','reachable', 'manipulator'], ascending=[False, False, False])
        print('file_list:',file_list[_])
        print('parts_list:',parts_list[_])
        best_reachable = df_sorted.iloc[0]['reachable']
        best_manipulator = df_sorted.iloc[0]['manipulator']
        best_torque = df_sorted.iloc[0]['torque']
        best_axis2 = df_sorted.iloc[0]['axis2']
        best_axis3 = df_sorted.iloc[0]['axis3']
        best_motor2 = df_sorted.iloc[0]['motor2']
        best_motor3 = df_sorted.iloc[0]['motor3']
        sheet_optimal_design.cell(row=i+1, column=1).value = parts_list[_]
        sheet_optimal_design.cell(row=i+1, column=2).value = best_torque
        sheet_optimal_design.cell(row=i+1, column=4).value = best_reachable
        sheet_optimal_design.cell(row=i+1, column=5).value = best_manipulator
        sheet_optimal_design.cell(row=i+1, column=6).value = best_axis2
        sheet_optimal_design.cell(row=i+1, column=7).value = best_axis3
        sheet_optimal_design.cell(row=i+1, column=8).value = best_motor2
        sheet_optimal_design.cell(row=i+1, column=9).value = best_motor3
        
        i = i + 1 
        # 與原設計比較
        f_ori = "tested_state_original_design_case1.xlsx"
        ori_df = pd.read_excel(f_ori,header=None, \
                        names=['mission_name', 'ratio','torque', 'reachable', 'manipulator', 'axis2', 'axis3',\
                                'motor2', 'motor3'])
        # 搜尋特定欄位中的值 mission_name# 選擇第一列
        mission_name_row = ori_df.iloc[:,0]

        # 搜尋第一列中特定的值
        search_value = parts_list[_]
        
        if search_value in mission_name_row.values:
                # 印出特定值所在的欄位位置
                index = list(mission_name_row.values).index(search_value)
                print(f'The search value "{search_value}" was found in column {index + 1}.')
                print(ori_df.iloc[index].transpose())

                # TODO: 比較
                # 访问最佳值的行和列
                best_reachable = df_sorted.iloc[0]['reachable']
                best_manipulator = df_sorted.iloc[0]['manipulator']
                best_torque = df_sorted.iloc[0]['torque']
                # 访问原設計值的行和列
                ori_reachable = ori_df.iloc[index]['reachable']
                ori_manipulator = ori_df.iloc[index]['manipulator']
                ori_torque = ori_df.iloc[index]['torque']

                reachable_diff = best_reachable - ori_reachable
                manipulator_diff = best_manipulator - ori_manipulator
                torque_diff = -best_torque + ori_torque

                print('reachable_diff: %.2f' % reachable_diff)
                print('manipulator_diff: %.5f' %  manipulator_diff)
                print('torque_diff: %.2f' %  torque_diff)
                total_reachable_diff += reachable_diff
                total_manipulator_diff += manipulator_diff
                total_torque_diff += torque_diff

                total_reachable_ori += ori_reachable
                total_manipulator_ori += ori_manipulator
                total_torque_ori += ori_torque

                total_reachable_best += best_reachable
                total_manipulator_best += best_manipulator
                total_torque_best += best_torque
                count_b += 1


        else:
                pass
                # print(f'The search value "{search_value}" was not found in the first row.')

file_name_optimal_design = "./tested_state_ddqn_optimal_design_case1_v3.xlsx"
optimal_design.save(file_name_optimal_design)


print('原設計平均可達性:%.2f' % (total_reachable_ori/count_a))
print('原設計平均可操控性:%.5f' % (total_manipulator_ori/count_b))
print('原設計平均torque:%.2f' % (total_torque_ori/(count_b)))

print('最佳平均可達性:%.2f' % (total_reachable_best/count_b))
print('最佳平均可操控性:%.5f' % (total_manipulator_best/count_b))
print('最佳平均torque:%.2f' % (total_torque_best/(count_b)))

print('total_reachable_diff: %.2f' % (total_reachable_diff/count_b))
print('total_manipulator_diff: %.5f' %  (total_manipulator_diff/count_b))
print('total_torque_diff: %.2f' %  (total_torque_diff/(count_b)))


