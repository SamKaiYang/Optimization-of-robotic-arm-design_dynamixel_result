import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, colors
from openpyxl.utils import get_column_letter
# 读取Excel文件
# df = pd.read_excel('C51-6-variable-20230508-193940/tested_reward_state_1-A-10.xlsx', header=None, \
                #    names=['ratio', 'torque', 'consuption', 'reachable', 'manipulator', 'axis2', 'axis3', 'all_length',\
                #            'nan', 'reward', 'nan', 'motor2', 'motor3'])
# df = pd.read_excel('test_result_excel/tested_reward_state_1-A-10.xlsx', header=None, \
                #    names=['ratio', 'torque', 'consuption', 'reachable', 'manipulator', 'axis2', 'axis3', 'all_length',\
                #            'nan', 'reward', 'nan', 'motor2', 'motor3'])
# 設定要讀取的檔案路徑和開頭字串
file_path = './0516/ddqn_tested_reward_state_0521'
file_prefix = 'tested_reward_state_'  # 或者是其他開頭字串 # 1-A-10

# 獲取符合開頭字串的所有檔案路徑
file_list = [os.path.join(file_path, f) for f in os.listdir(file_path) if f.startswith(file_prefix)]
print('file_list:',file_list)

# Use list comprehension to extract the parts from the file names
parts_list = [file_name.split('/')[-1].split('.')[0].split('_')[-1] for file_name in file_list]

print(parts_list)

df_list = [pd.read_excel(f,header=None, \
                   names=['ratio', 'torque', 'consuption', 'reachable', 'manipulator', 'axis2', 'axis3', 'all_length',\
                           'nan', 'reward', 'nan', 'motor2', 'motor3']) for f in file_list]

total_reachable_diff = 0
total_manipulator_diff = 0
total_power_consumption_diff = 0
total_ratio_diff = 0
total_torque_diff = 0

total_reachable_ori = 0
total_manipulator_ori = 0
total_power_consumption_ori = 0
total_ratio_ori = 0
total_torque_ori = 0

total_reachable_best = 0
total_manipulator_best = 0
total_power_consumption_best = 0
total_ratio_best = 0
total_torque_best = 0

optimal_design = Workbook()
sheet_optimal_design = optimal_design.active
i = 0

count_a = 0
count_b = 0
for _ in range(len(file_list)):
        df_sorted = df_list[_].sort_values(by=['reward','reachable', 'manipulator', 'consuption'], ascending=[False, False, False, True])
        print('file_list:',file_list[_])
        print('parts_list:',parts_list[_])
        # 打印排序后的DataFrame，以确认最佳值
        print(df_sorted)

        # 访问最佳值的行和列
        print(df_sorted.iloc[0].transpose())
        best_reachable = df_sorted.iloc[0]['reachable']
        best_manipulator = df_sorted.iloc[0]['manipulator']
        best_power_consumption = df_sorted.iloc[0]['consuption']
        best_ratio = df_sorted.iloc[0]['ratio']
        best_torque = df_sorted.iloc[0]['torque']
        best_axis2 = df_sorted.iloc[0]['axis2']
        best_axis3 = df_sorted.iloc[0]['axis3']
        best_motor2 = df_sorted.iloc[0]['motor2']
        best_motor3 = df_sorted.iloc[0]['motor3']
        sheet_optimal_design.cell(row=i+1, column=1).value = parts_list[_]
        sheet_optimal_design.cell(row=i+1, column=2).value = best_ratio
        sheet_optimal_design.cell(row=i+1, column=3).value = best_torque
        sheet_optimal_design.cell(row=i+1, column=4).value = best_power_consumption
        sheet_optimal_design.cell(row=i+1, column=5).value = best_reachable
        sheet_optimal_design.cell(row=i+1, column=6).value = best_manipulator
        sheet_optimal_design.cell(row=i+1, column=7).value = best_axis2
        sheet_optimal_design.cell(row=i+1, column=8).value = best_axis3
        sheet_optimal_design.cell(row=i+1, column=9).value = best_motor2
        sheet_optimal_design.cell(row=i+1, column=10).value = best_motor3
        
        i = i + 1 
        # 與原設計比較
        f_ori = "tested_state_original_design.xlsx"
        ori_df = pd.read_excel(f_ori,header=None, \
                        names=['mission_name', 'ratio','torque', 'consuption', 'reachable', 'manipulator', 'axis2', 'axis3',\
                                'motor2', 'motor3'])
        # 搜尋特定欄位中的值 mission_name# 選擇第一列
        mission_name_row = ori_df.iloc[:,0]

        # 搜尋第一列中特定的值
        search_value = parts_list[_]
        
        if search_value in mission_name_row.values:
                # 印出特定值所在的欄位位置
                index = list(mission_name_row.values).index(search_value)
                print(f'The search value "{search_value}" was found in column {index + 1}.')
                print(ori_df.iloc[index].transpose())

                # TODO: 比較
                # 访问最佳值的行和列
                best_reachable = df_sorted.iloc[0]['reachable']
                best_manipulator = df_sorted.iloc[0]['manipulator']
                best_power_consumption = df_sorted.iloc[0]['consuption']
                best_ratio = df_sorted.iloc[0]['ratio']
                best_torque = df_sorted.iloc[0]['torque']
                # 访问原設計值的行和列
                ori_reachable = ori_df.iloc[index]['reachable']
                print("ori_reachable:",ori_reachable)
                ori_manipulator = ori_df.iloc[index]['manipulator']
                ori_power_consumption = ori_df.iloc[index]['consuption']
                ori_ratio = ori_df.iloc[index]['ratio']
                ori_torque = ori_df.iloc[index]['torque']

                if int(ori_reachable) != 1:
                        print("僅比較可達性與可操控性")
                        reachable_diff = best_reachable - ori_reachable
                        manipulator_diff = best_manipulator - ori_manipulator
                        print('reachable_diff: %.2f' % reachable_diff)
                        print('manipulator_diff: %.5f' %  manipulator_diff)
                        # print('power_consumption_diff: %.2f' %  power_consumption_diff)
                        # print('ratio_diff: %.2f' %  ratio_diff)
                        # print('torque_diff: %.2f' %  torque_diff)
                        total_reachable_diff += reachable_diff
                        total_manipulator_diff += manipulator_diff
                        total_reachable_ori += ori_reachable
                        total_manipulator_ori += ori_manipulator
                        total_reachable_best += best_reachable
                        total_manipulator_best += best_manipulator
                        count_a += 1
                else:
                        print("比較全部")
                        reachable_diff = best_reachable - ori_reachable
                        manipulator_diff = best_manipulator - ori_manipulator
                        power_consumption_diff = -best_power_consumption + ori_power_consumption
                        ratio_diff = -best_ratio + ori_ratio
                        torque_diff = -best_torque + ori_torque

                        print('reachable_diff: %.2f' % reachable_diff)
                        print('manipulator_diff: %.5f' %  manipulator_diff)
                        print('power_consumption_diff: %.2f' %  power_consumption_diff)
                        print('ratio_diff: %.2f' %  ratio_diff)
                        print('torque_diff: %.2f' %  torque_diff)
                        total_reachable_diff += reachable_diff
                        total_manipulator_diff += manipulator_diff
                        total_power_consumption_diff += power_consumption_diff
                        total_ratio_diff += ratio_diff
                        total_torque_diff += torque_diff

                        total_reachable_ori += ori_reachable
                        total_manipulator_ori += ori_manipulator
                        total_power_consumption_ori += ori_power_consumption
                        total_ratio_ori += ori_ratio
                        total_torque_ori += ori_torque

                        total_reachable_best += best_reachable
                        total_manipulator_best += best_manipulator
                        total_power_consumption_best += best_power_consumption
                        total_ratio_best += best_ratio
                        total_torque_best += best_torque
                        count_b += 1


        else:
                print(f'The search value "{search_value}" was not found in the first row.')

file_name_optimal_design = "./tested_state_c51_optimal_design.xlsx"
optimal_design.save(file_name_optimal_design)


print('原設計平均可達性:%.2f' % (total_reachable_ori/(count_a+count_b)))
print('原設計平均可操控性:%.5f' % (total_manipulator_ori/(count_a+count_b)))
print('原設計平均功耗:%.2f' % (total_power_consumption_ori/(count_b)))
print('原設計平均ratio:%.2f' % (total_ratio_ori/(count_b)))
print('原設計平均torque:%.2f' % (total_torque_ori/(count_b)))

print('最佳平均可達性:%.2f' % (total_reachable_best/(count_a+count_b)))
print('最佳平均可操控性:%.5f' % (total_manipulator_best/(count_a+count_b)))
print('最佳平均功耗:%.2f' % (total_power_consumption_best/(count_b)))
print('最佳平均ratio:%.2f' % (total_ratio_best/(count_b)))
print('最佳平均torque:%.2f' % (total_torque_best/(count_b)))

print('total_reachable_diff: %.2f' % (total_reachable_diff/(count_a+count_b)))
print('total_manipulator_diff: %.5f' %  (total_manipulator_diff/(count_a+count_b)))
print('total_power_consumption_diff: %.2f' %  (total_power_consumption_diff/(count_b)))
print('total_ratio_diff: %.2f' %  (total_ratio_diff/(count_b)))
print('total_torque_diff: %.2f' %  (total_torque_diff/(count_b)))


