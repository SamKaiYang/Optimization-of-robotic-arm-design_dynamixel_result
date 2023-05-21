# import pandas as pd
# import matplotlib.pyplot as plt

# # 讀取Excel資料到DataFrame
# df = pd.read_excel('test_result_excel/tested_state_traj_1_3-A-10.xlsx', header=None)

# # 設置x軸和y軸數據
# x_data = df.index.values
# y_data = df.iloc[:, 2].values

# # 繪製變化圖
# plt.plot(x_data, y_data)
# plt.xlabel('Index')
# plt.ylabel('Data')
# plt.show()


from openpyxl import load_workbook
import matplotlib.pyplot as plt
import itertools
import yaml
import os
###　test 4
todo_show_flag = False
data = '0514'
case_name = '/ddqn_tested_state_traj'
# 讀取Excel檔案
with open('muti_mission_input.yaml', 'r') as f:
        config = yaml.load(f, Loader=yaml.Loader)
payload = config['payload']
work_space = config['work_space']
mission_time = config['mission_time']
combinations = list(itertools.product(payload, work_space, mission_time))
traj_num = 0

for p, w, t in itertools.product(payload, work_space, mission_time):
    switch_mission = f'{p}-{w}-{t}'
    for traj_num in range(100):
        try:
            # 嘗試讀取檔案
            # traj_num = traj_num + 1 
            filename= data + case_name + '/tested_state_traj_' + str(traj_num) + '_' + switch_mission + '.xlsx'
            print(filename)
            workbook = load_workbook(filename)
            worksheet = workbook.active
            i = 0
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            plt.plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            plt.title('speed limit')
            plt.xlabel('times')
            plt.ylabel('counts')
            # 顯示圖表
            if todo_show_flag is not False:
                plt.show()

            i = 1
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            plt.plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            plt.title('torque limit')
            plt.xlabel('times')
            plt.ylabel('counts')
            # 顯示圖表
            if todo_show_flag is not False:
                plt.show()

            i = 2
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            plt.plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            plt.title('power consumption')
            plt.xlabel('times')
            plt.ylabel('J')
            # 顯示圖表
            if todo_show_flag is not False:
                plt.show()

            i = 3
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            plt.plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            plt.title('reachability')
            plt.xlabel('times')
            plt.ylabel('score')
            # 顯示圖表
            if todo_show_flag is not False:
                plt.show()
            i = 4
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            plt.plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            plt.title('manipulability')
            plt.xlabel('times')
            plt.ylabel('index')
            # 顯示圖表
            if todo_show_flag is not False:
                plt.show()


            fig, ax = plt.subplots(nrows=5, ncols=1, figsize=(6, 18), sharex=True)
            i = 0
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            ax[i].plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            ax[i].set_title('speed limit')
            ax[i].set_xlabel('times')
            ax[i].set_ylabel('counts')
            ax[i].set_xticks([0,5,10,15,20,25,30,35,40,45,50])
            # ax[i].set_aspect('auto')
            # ## 於軸上顯示較小的刻度
            # ax[i].minorticks_on()
            # ax[i].set_xticklabels(
            #         labels=labels,
            #         fontsize=5,
            #         rotation=0)
            i = 1
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            ax[i].plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            ax[i].set_title('torque limit')
            ax[i].set_xlabel('times')
            ax[i].set_ylabel('counts')


            i = 2
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            ax[i].plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            ax[i].set_title('power consumption')
            ax[i].set_xlabel('times')
            ax[i].set_ylabel('W')

            i = 3
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            ax[i].plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            ax[i].set_title('reachability')
            ax[i].set_xlabel('times')
            ax[i].set_ylabel('score')

            i = 4
            col_data = list(worksheet.iter_cols(min_row=1, max_row=46, min_col=i+1, max_col=i+1, values_only=True))[0]
            # 获取 x 坐标和 y 坐标
            # x = len(data[:]) # 46 筆資料
            x = range(1, len(col_data)+1)
            y = col_data
            # 绘制折线图
            ax[i].plot(x, y)

            # 添加标题和轴标签
            # plt.title('Column {}'.format(i))
            ax[i].set_title('manipulability')
            ax[i].set_xlabel('times')
            ax[i].set_ylabel('index')
            if not os.path.exists('traj_plot_result'+ case_name):
                        os.makedirs('traj_plot_result'+ case_name)
            save_picture_name = 'traj_plot_result'+ case_name +'/plot_'+ str(traj_num) + '_' + switch_mission +'.png'
            print(save_picture_name)
            plt.savefig(save_picture_name) #保存图片
            # 顯示圖表
            if todo_show_flag is not False:
                plt.show()
        except FileNotFoundError:
            # 當找不到檔案時的處理邏輯
            print(f"無法找到檔案: {filename}")
        except Exception as e:
            # 其他異常的處理邏輯
            print(f"處理檔案時發生錯誤: {e}")
